import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.load_workbook('GGESTION DE RIESGOS - ACTIVOS DE INF-2026.xlsx', data_only=False)

# Delete existing sheet if it was created accidentally
if 'ANÁLISIS_CRITICIDAD_RIESGO' in wb.sheetnames:
    del wb['ANÁLISIS_CRITICIDAD_RIESGO']

inv_ws = wb['INVENTARIO']

# Find max row based on column A (ID)
actual_max_row = 2
while inv_ws.cell(row=actual_max_row+1, column=1).value is not None:
    actual_max_row += 1

print(f"Max row in INVENTARIO: {actual_max_row}")

# Nivel de Madurez (Col AA): L0 Inexistente, L1 Inicial / Ad Hoc, L2 Reproducible pero Intuitivo, L3 Proceso Definido, L4 Gestionado y Medible, L5 Optimizado
# Estado de Implementación (Col Y): Implementado, En proceso, No Iniciado, No Aplica
# Tipo de Protección (Col Z): PREVENTIVO, DETECTIVO, ADMINISTRACION, RECUPERACION, ELIMINACION
# Riesgos / Amenazas (Col AB) - we will provide a comma separated list
# Evidencia (Col AC) - we will provide a comma separated list

madurez_list = '"L0 Inexistente,L1 Inicial / Ad Hoc,L2 Reproducible pero Intuitivo,L3 Proceso Definido,L4 Gestionado y Medible,L5 Optimizado"'
estado_list = '"Implementado,En proceso,No Iniciado,No Aplica"'
proteccion_list = '"PREVENTIVO,DETECTIVO,ADMINISTRACION,RECUPERACION,ELIMINACION"'
riesgos_list = '"[A.5] Suplantación de la identidad del usuario,[A.6] Abuso de privilegios de acceso,[A.7] Uso no previsto,[A.11] Acceso no autorizado,[A.19] Divulgación de información,[A.24] Denegación de servicio,[A.25] Robo,[A.30] Ingeniería social (Phishing),[E.2] Errores del administrador,[E.18] Destrucción de información,[E.19] Fugas de información,[E.20] Vulnerabilidades de los programas (software),[E.23] Errores de mantenimiento / actualización de equipos (hardware),[E.24] Caída del sistema por agotamiento de recursos,[E.25] Pérdida de equipos,[E.28] Indisponibilidad del personal,[N.1] Fuego,[N.2] Daños por agua"'
evidencia_list = '"Manual de Políticas de Seguridad de la Información (MPSI),Manual de Gestión de Seguridad de la Información (MGSI),Matriz de Roles y Perfiles,Informe Técnico de Pentesting / Ethical Hacking,Bitácora de Seguridad de la Información,Contrato de Proveedor / Terceros,Acta de Respaldos y Resguardo"'

dv_estado = DataValidation(type="list", formula1=estado_list, allow_blank=True)
dv_proteccion = DataValidation(type="list", formula1=proteccion_list, allow_blank=True)
dv_madurez = DataValidation(type="list", formula1=madurez_list, allow_blank=True)
dv_riesgos = DataValidation(type="list", formula1=riesgos_list, allow_blank=True)
dv_evidencia = DataValidation(type="list", formula1=evidencia_list, allow_blank=True)

inv_ws.add_data_validation(dv_estado)
inv_ws.add_data_validation(dv_proteccion)
inv_ws.add_data_validation(dv_madurez)
inv_ws.add_data_validation(dv_riesgos)
inv_ws.add_data_validation(dv_evidencia)

# Apply validations to respective columns
for row in range(3, actual_max_row + 1):
    dv_estado.add(inv_ws.cell(row=row, column=25))   # Col Y
    dv_proteccion.add(inv_ws.cell(row=row, column=26)) # Col Z
    dv_madurez.add(inv_ws.cell(row=row, column=27))  # Col AA
    dv_riesgos.add(inv_ws.cell(row=row, column=28))  # Col AB
    dv_evidencia.add(inv_ws.cell(row=row, column=29)) # Col AC

# Header for column AD (col 30)
header_col = 30
inv_ws.cell(row=2, column=header_col, value="Evaluación de Cumplimiento")
inv_ws.cell(row=2, column=header_col).font = Font(bold=True)

for row in range(3, actual_max_row + 1):
    crit = f'T{row}'
    estado = f'Y{row}'
    mad = f'AA{row}'
    evid = f'AC{row}'

    # We use exact strings to be extremely precise
    # Critical condition: Criticidad is ALTA AND (Madurez is L0, L1, or Estado is No Iniciado)
    is_critico_cond = f'OR({mad}="L0 Inexistente",{mad}="L1 Inicial / Ad Hoc",{estado}="No Iniciado")'
    cond1 = f'AND({crit}="ALTA",{is_critico_cond})'
    res1 = '"CRÍTICO - Incumplimiento Normativo SEPS"'

    # Conforme condition: Madurez is L4 or L5 AND Evidencia is not blank and not "N/A" (or we just check if it's not blank)
    is_madurez_alta = f'OR({mad}="L4 Gestionado y Medible",{mad}="L5 Optimizado")'
    has_evidencia = f'AND({evid}<>"",{evid}<>"N/A")'
    cond2 = f'AND({is_madurez_alta},{has_evidencia})'
    res2 = '"CONFORME - Control Efectivo ISO 27001"'

    res3 = '"OBSERVACIÓN - Revisar Controles"'

    # Only show if not blank
    eval_formula = f'=IF(OR({crit}="",{estado}="",{mad}=""),"",IF({cond1},{res1},IF({cond2},{res2},{res3})))'

    inv_ws.cell(row=row, column=header_col, value=eval_formula)

red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

range_str = f'AD3:AD{actual_max_row}'

inv_ws.conditional_formatting.add(
    range_str,
    CellIsRule(operator='equal', formula=['"CRÍTICO - Incumplimiento Normativo SEPS"'], stopIfTrue=False, fill=red_fill)
)

inv_ws.conditional_formatting.add(
    range_str,
    CellIsRule(operator='equal', formula=['"CONFORME - Control Efectivo ISO 27001"'], stopIfTrue=False, fill=green_fill)
)

inv_ws.conditional_formatting.add(
    range_str,
    CellIsRule(operator='equal', formula=['"OBSERVACIÓN - Revisar Controles"'], stopIfTrue=False, fill=yellow_fill)
)

# Optional: Add some dummy values to Row 3 for testing logic visibly for the user
# inv_ws.cell(row=3, column=25, value="No Iniciado")
# inv_ws.cell(row=3, column=27, value="L0 Inexistente")
# Wait, let's not touch their actual data, they can select it via dropdown.

wb.save('GGESTION DE RIESGOS - ACTIVOS DE INF-2026.xlsx')
print("Successfully saved GGESTION DE RIESGOS - ACTIVOS DE INF-2026.xlsx")
