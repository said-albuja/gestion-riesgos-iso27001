import openpyxl
wb = openpyxl.load_workbook('GGESTION DE RIESGOS - ACTIVOS DE INF-2026.xlsx', data_only=False)
ws = wb['INVENTARIO']
assert 'ANÁLISIS_CRITICIDAD_RIESGO' not in wb.sheetnames, "Should have been deleted"
assert ws['AD2'].value == 'Evaluación de Cumplimiento', "Header incorrect"
assert 'CRÍTICO' in ws['AD1000'].value, "Formula looks off"
assert ws.max_row >= 1406
print("All checks passed.")
