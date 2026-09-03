import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('GGESTION DE RIESGOS - ACTIVOS DE INF-2026.xlsx', data_only=False)

# Sanitize validations to prevent crashes
for ws in wb.worksheets:
    ws.data_validations.dataValidation = []

if 'ANÁLISIS_CRITICIDAD_RIESGO' in wb.sheetnames:
    del wb['ANÁLISIS_CRITICIDAD_RIESGO']

inv_ws = wb['INVENTARIO']
max_row = inv_ws.max_row

# Find actual max row (stop when ID is empty)
actual_max_row = 2
while inv_ws.cell(row=actual_max_row+1, column=1).value is not None:
    actual_max_row += 1

print(f"Max row in INVENTARIO: {actual_max_row}")

# Add header to Column AD (30) on row 2 (which is where headers seem to be)
header_col = 30
inv_ws.cell(row=2, column=header_col, value="Evaluación de Cumplimiento")
inv_ws.cell(row=2, column=header_col).font = Font(bold=True)

# The user wants formulas to be "understandable".
# CRITICIDAD is in column T.
# Estado de Implementación is in column Y.
# Nivel de Madurez / Efectividad is in column AA.
# Evidencia / Documentación Asociada is in column AC.

for row in range(3, actual_max_row + 1):
    crit = f'T{row}'
    estado = f'Y{row}'
    mad = f'AA{row}'
    evid = f'AC{row}'

    # "CRÍTICO - Incumplimiento Normativo SEPS":
    # IF Criticidad = "ALTA" AND (Madurez = "L0*" or Madurez = "L1*" or Estado = "No Iniciado")

    # "CONFORME - Control Efectivo ISO 27001":
    # IF (Madurez = "L4*" or Madurez = "L5*") AND Evidencia is not empty and not "N/A"

    # "OBSERVACIÓN - Revisar Controles": otherwise

    # Understandable formula in Spanish Excel (using commas, as openpyxl translates them):
    is_critico_cond = f'OR(LEFT({mad},2)="L0",LEFT({mad},2)="L1",{estado}="No Iniciado")'
    cond1 = f'AND({crit}="ALTA",{is_critico_cond})'
    res1 = '"CRÍTICO - Incumplimiento Normativo SEPS"'

    is_madurez_alta = f'OR(LEFT({mad},2)="L4",LEFT({mad},2)="L5")'
    has_evidencia = f'AND({evid}<>"",{evid}<>"N/A")'
    cond2 = f'AND({is_madurez_alta},{has_evidencia})'
    res2 = '"CONFORME - Control Efectivo ISO 27001"'

    res3 = '"OBSERVACIÓN - Revisar Controles"'

    eval_formula = f'=IF(AND({crit}="",{mad}="",{estado}=""),"",IF({cond1},{res1},IF({cond2},{res2},{res3})))'

    inv_ws.cell(row=row, column=header_col, value=eval_formula)

red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

range_str = f'AD3:AD{actual_max_row}'

inv_ws.conditional_formatting.add(
    range_str,
    CellIsRule(operator='equal', formula=['"CRÍTICO - Incumplimiento Normativo SEPS"'], stopIfTrue=True, fill=red_fill)
)

inv_ws.conditional_formatting.add(
    range_str,
    CellIsRule(operator='equal', formula=['"CONFORME - Control Efectivo ISO 27001"'], stopIfTrue=True, fill=green_fill)
)

inv_ws.conditional_formatting.add(
    range_str,
    CellIsRule(operator='equal', formula=['"OBSERVACIÓN - Revisar Controles"'], stopIfTrue=True, fill=yellow_fill)
)

wb.save('GGESTION DE RIESGOS - ACTIVOS DE INF-2026.xlsx')
print("Successfully saved GGESTION DE RIESGOS - ACTIVOS DE INF-2026.xlsx")
